import decimal
import requests
from datetime import date
from bs4 import BeautifulSoup
from openpyxl import load_workbook


def main():
	wb = load_workbook('data.xlsx')
	sheet = wb.active

	today = date.today()
	currentDate = today.strftime("%m/%d/%Y")

	r, c = 2, 5

	while sheet.cell(row = r, column = c).value != None:
		c += 1

	if sheet.cell(row = 1, column = c - 1).value == currentDate:
		c -= 1
	else:
		sheet.cell(row = 1, column = c).value = currentDate
	
	while sheet.cell(row = r, column = 1).value != None:	
		url = sheet.cell(row = r, column = 1).value
		item = Item(r, url)

		print('Sending request for row number ', r, end = '\r')
		item.requestToAmazon()

		if sheet.cell(row = r, column = 2).value == None or sheet.cell(row = r, column = 2).value == 'Failed to read':
			sheet.cell(row = r, column = 2).value = item.asin

		if sheet.cell(row = r, column = 3).value == None or sheet.cell(row = r, column = 3).value == 'Failed to read':
			sheet.cell(row = r, column = 3).value = item.title

		if item.price != 'Failed to read':
			sheet.cell(row = r, column = 4).value = item.price
			sheet.cell(row = r, column = 4).style = 'Currency'
			sheet.cell(row = r, column = c).value = item.price
			sheet.cell(row = r, column = c).style = 'Currency'

		print('Request for row number ', r, ' complete')
		r += 1
	
	wb.save("data.xlsx")


class Item:
	def __init__(self, row, url):
		self.row = row
		self.url = url


	def requestToAmazon(self):
		HEADERS = ({'User-Agent':
			'Mozilla/5.0 (Windows NT 6.1) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/41.0.2228.0 Safari/537.36',
			'Accept-Language': 'en-US, en;q=0.5'})

		r = requests.get(self.url, headers=HEADERS)
		soup = BeautifulSoup(r.text, 'lxml')
        
		try:
			self.title = soup.find(id='productTitle').get_text().strip()
		except:
			self.title = 'Failed to read'
			print('*** Failed to read title ***')

		try:
			priceDiv = soup.find(id='corePrice_feature_div')
			itemPrice = priceDiv.find_all("span", {"class": "a-offscreen"})[0].text
			self.price = decimal.Decimal(itemPrice[1:])
		except:
			self.price = "Failed to read"
			print('*** Failed to read price ***')

		try:
			asinDiv = soup.find(id='addToCart')
			self.asin = asinDiv.find(id='ASIN').get('value')
		except:
			self.asin = "Failed to read"
			print('*** Failed to read asin ***')


if __name__ == '__main__':
	main()

